"""
ECI Candidate Affidavit Scraper
URL: https://affidavit.eci.gov.in/candidate-affidavit

Flow:
  1. Load page with Playwright (headless=False — Akamai blocks headless)
  2. Select election → state → phase → constituency → Filter
  3. For each candidate → click "View more" → click Download
  4. Intercept /affidavit-pdf-download/ URL → download PDF
  5. Deduplicate by nomid+candidateid (stable IDs from /increaseDownloadCount)

Deduplication:
  - .downloaded.json per constituency tracks {nomid_candidateid: filename}
  - Re-run safely — only new affidavits are downloaded

Output: eci_affidavit_downloads/{constituency_name}/Affidavit-*.pdf

Run:
  python3 src/scrapers/implementations/eci_affidavit/affidavit_scraper.py
  python3 src/scrapers/implementations/eci_affidavit/affidavit_scraper.py --constituency TIRUTTANI
"""
import asyncio
import json
import logging
from datetime import datetime
from pathlib import Path
import re
import argparse
from pathlib import Path

from playwright.async_api import async_playwright, Page, BrowserContext

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("eci_affidavit")

BASE_URL = "https://affidavit.eci.gov.in"
DOWNLOAD_DIR = Path("eci_affidavit_downloads")
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36"


# ─────────────────────────────────────────────────────────────────────────────
# Deduplication
# ─────────────────────────────────────────────────────────────────────────────

def load_tracker(path: Path) -> dict:
    """Returns {nomid_candidateid: filename}"""
    if path.exists():
        try:
            return json.loads(path.read_text()).get("downloaded", {})
        except Exception:
            return {}
    return {}


def update_daily_report(filter_date: str, constituency: str, downloaded_this_run: int, total_downloaded: int):
    """Append/update daily Excel report."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime as _dt

        date_str = filter_date or _dt.now().strftime("%d-%m-%Y")
        report_path = DOWNLOAD_DIR / f"report_{date_str}.xlsx"

        # Load or create workbook
        if report_path.exists():
            wb = openpyxl.load_workbook(report_path)
        else:
            wb = openpyxl.Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        sheet_name = date_str
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            # Header row
            headers = ["Constituency", "Downloaded This Run", "Total Downloaded", "Last Run Time"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E79")
                cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 20
        else:
            ws = wb[sheet_name]

        # Find existing row for this constituency or add new
        run_time = _dt.now().strftime("%d-%m-%Y %H:%M")
        found = False
        for row in ws.iter_rows(min_row=2):
            if row[0].value == constituency:
                row[1].value = downloaded_this_run
                row[2].value = total_downloaded
                row[3].value = run_time
                found = True
                break

        if not found:
            ws.append([constituency, downloaded_this_run, total_downloaded, run_time])

        wb.save(report_path)
    except Exception as e:
        logger.warning(f"Report update failed: {e}")


def load_tracker(path: Path) -> dict:
    """Returns {candidateid|upload_time: filename}"""
    if path.exists():
        try:
            return json.loads(path.read_text()).get("downloaded", {})
        except Exception:
            return {}
    return {}


def save_tracker(path: Path, downloaded: dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps({"downloaded": downloaded}, indent=2))


# ─────────────────────────────────────────────────────────────────────────────
# Browser setup
# ─────────────────────────────────────────────────────────────────────────────

async def make_browser(playwright):
    browser = await playwright.chromium.launch(
        headless=False,
        args=["--disable-blink-features=AutomationControlled"],
    )
    context = await browser.new_context(
        user_agent=UA,
        viewport={"width": 1280, "height": 800},
        accept_downloads=True,
    )
    await context.add_init_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    return browser, context


async def select_if_exists(page: Page, selector: str, value: str):
    try:
        await page.select_option(selector, value=value, timeout=5000)
    except Exception:
        pass


async def get_options(page: Page, selector: str) -> list[tuple[str, str]]:
    opts = await page.query_selector_all(f"{selector} option")
    result = []
    for o in opts:
        val = await o.get_attribute("value") or ""
        txt = (await o.inner_text()).strip()
        if val:
            result.append((val, txt))
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Scraper
# ─────────────────────────────────────────────────────────────────────────────

class AffidavitScraper:

    def __init__(self, state_code: str, election_type: str, target_constituency=None,
                 upload: bool = False, filter_date: str = None, before_time: str = None):
        self.state_code = state_code
        self.election_type = election_type
        # Accept single string or list
        if isinstance(target_constituency, list):
            self.target_constituencies = [c.upper() for c in target_constituency]
        elif target_constituency:
            self.target_constituencies = [target_constituency.upper()]
        else:
            self.target_constituencies = []
        self.upload = upload
        self.filter_date = filter_date
        self.before_time = before_time
        self._upload_client = None  # optional filter

    async def run(self):
        DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
        checkpoint_file = DOWNLOAD_DIR / ".run_checkpoint.json"

        def load_checkpoint():
            if checkpoint_file.exists():
                try:
                    return json.loads(checkpoint_file.read_text())
                except Exception:
                    pass
            return {"completed": [], "last_const": None}

        def save_checkpoint(completed: list, last_const: str):
            checkpoint_file.write_text(json.dumps({
                "completed": completed,
                "last_const": last_const,
                "date_filter": self.filter_date,
                "updated_at": str(datetime.now()),
            }, indent=2))

        checkpoint = load_checkpoint()
        completed_constituencies = set(checkpoint.get("completed", []))
        # Checkpoint only applies to full runs (no --constituency filter)
        # For specific constituency runs, always start fresh
        if self.target_constituencies:
            completed_constituencies = set()
            if checkpoint_file.exists():
                checkpoint_file.unlink()
        else:
            # For full runs, only resume if checkpoint is recent (within 2 hours = same crash session)
            checkpoint_age_hours = 999
            if checkpoint_file.exists():
                try:
                    import time as _time
                    age_seconds = _time.time() - checkpoint_file.stat().st_mtime
                    checkpoint_age_hours = age_seconds / 3600
                except Exception:
                    pass
            if checkpoint_age_hours > 2:
                completed_constituencies = set()
                if checkpoint_file.exists():
                    checkpoint_file.unlink()
                    logger.info("Starting fresh full run (previous checkpoint expired)")
            elif completed_constituencies:
                logger.info(f"Resuming from crash — {len(completed_constituencies)} constituencies already done")
        async with async_playwright() as p:
            browser, context = await make_browser(p)
            page = await context.new_page()

            try:
                logger.info("Loading ECI affidavit portal...")
                # Init upload client if needed
                if self.upload:
                    import importlib.util, os
                    spec = importlib.util.spec_from_file_location(
                        "upload_client",
                        os.path.join(os.path.dirname(__file__), "upload_client.py")
                    )
                    mod = importlib.util.module_from_spec(spec)
                    spec.loader.exec_module(mod)
                    self._upload_client = mod.UploadClient()
                    await self._upload_client.init()
                # Load page with retry (ECI site can be flaky)
                for attempt in range(3):
                    try:
                        await page.goto(f"{BASE_URL}/candidate-affidavit", wait_until="domcontentloaded", timeout=60000)
                        break
                    except Exception as e:
                        if attempt == 2:
                            raise
                        logger.warning(f"Page load failed (attempt {attempt+1}), retrying in 5s... {e}")
                        await asyncio.sleep(5)
                logger.info(f"Page: {await page.title()}")

                # Select election type
                await page.select_option("select[name=electionType]", value=self.election_type)
                await page.wait_for_timeout(800)

                # Select first election sub-type (AC - GENERAL)
                election_opts = await get_options(page, "select[name=election]")
                if election_opts:
                    await page.select_option("select[name=election]", value=election_opts[0][0])
                    await page.wait_for_timeout(500)

                # Select state
                await page.select_option("select[name=states]", value=self.state_code)
                await page.wait_for_timeout(2500)

                # Get phases — retry up to 5s for AJAX to populate
                phase_opts = []
                for _ in range(10):
                    phase_opts = await get_options(page, "select[name=phase]")
                    if phase_opts:
                        break
                    await page.wait_for_timeout(500)
                logger.info(f"Phases: {phase_opts}")

                if not phase_opts:
                    phase_opts = [("", "All")]

                for phase_val, phase_name in phase_opts:
                    if phase_val:
                        await select_if_exists(page, "select[name=phase]", phase_val)
                        await page.wait_for_timeout(1000)

                    # Wait for constituency dropdown to populate (AJAX call)
                    for _ in range(10):
                        const_opts = await get_options(page, "select[name=constId]")
                        if const_opts:
                            break
                        await page.wait_for_timeout(500)
                    logger.info(f"Phase {phase_name}: {len(const_opts)} constituencies")

                    for const_val, const_name in const_opts:
                        if self.target_constituencies and not any(
                            t in const_name.upper() for t in self.target_constituencies
                        ):
                            continue
                        # Skip already completed constituencies (resume support)
                        if const_name in completed_constituencies:
                            logger.info(f"  Skipping {const_name} — already completed")
                            # Still update report for skipped ones
                            tracker_file = DOWNLOAD_DIR / re.sub(r'[^\w\s-]', '', const_name).strip().replace(' ', '_') / ".downloaded.json"
                            raw = load_tracker(tracker_file)
                            total_dl = sum(1 for k in raw if "|" not in str(k))
                            update_daily_report(self.filter_date, const_name, 0, total_dl)
                            continue
                        downloaded_this_run = await self._process_constituency(page, context, phase_val, const_val, const_name)
                        # Save checkpoint after each constituency
                        completed_constituencies.add(const_name)
                        save_checkpoint(list(completed_constituencies), const_name)
                        logger.info(f"  ✓ Checkpoint saved ({len(completed_constituencies)} done)")
                        # Update daily Excel report — count only candidateid entries (not version_key entries)
                        tracker_file = DOWNLOAD_DIR / re.sub(r'[^\w\s-]', '', const_name).strip().replace(' ', '_') / ".downloaded.json"
                        raw = load_tracker(tracker_file)
                        # candidateid entries are plain numbers or short strings without "|"
                        total_dl = sum(1 for k in raw if "|" not in str(k))
                        update_daily_report(self.filter_date, const_name, downloaded_this_run, total_dl)

            finally:
                await browser.close()
        # Clear checkpoint on successful completion
        if checkpoint_file.exists() and not self.target_constituencies:
            checkpoint_file.unlink()
            logger.info("All constituencies completed — checkpoint cleared")

    async def _process_constituency(self, page: Page, context: BrowserContext,
                                     phase_val: str, const_val: str, const_name: str):
        safe_name = re.sub(r'[^\w\s-]', '', const_name).strip().replace(' ', '_')
        out_dir = DOWNLOAD_DIR / safe_name
        tracker_file = out_dir / ".downloaded.json"
        downloaded = load_tracker(tracker_file)

        logger.info(f"Constituency: {const_name} | already downloaded: {len(downloaded)}")

        # Navigate back to filter page and reselect
        for nav_attempt in range(3):
            try:
                await page.goto(f"{BASE_URL}/candidate-affidavit", wait_until="domcontentloaded", timeout=60000)
                await page.wait_for_timeout(500)
                await page.select_option("select[name=electionType]", value=self.election_type)
                await page.wait_for_timeout(800)
                election_opts = await get_options(page, "select[name=election]")
                if election_opts:
                    await page.select_option("select[name=election]", value=election_opts[0][0])
                    await page.wait_for_timeout(600)

                # Wait for states dropdown to populate before selecting
                for _ in range(15):
                    state_opts = await get_options(page, "select[name=states]")
                    if any(v == self.state_code for v, _ in state_opts):
                        break
                    await page.wait_for_timeout(500)
                await page.select_option("select[name=states]", value=self.state_code)
                await page.wait_for_timeout(1500)
                break  # success
            except Exception as e:
                if nav_attempt == 2:
                    raise
                logger.warning(f"Navigation retry {nav_attempt+1}: {e}")
                await page.wait_for_timeout(3000)

        if phase_val:
            await select_if_exists(page, "select[name=phase]", phase_val)
            await page.wait_for_timeout(1000)

        # Wait for constituency dropdown to populate
        for _ in range(10):
            const_check = await get_options(page, "select[name=constId]")
            if const_check:
                break
            await page.wait_for_timeout(500)

        await select_if_exists(page, "select[name=constId]", const_val)
        await page.wait_for_timeout(300)

        # Submit filter
        await page.click("button[type=submit]")
        await page.wait_for_load_state("domcontentloaded", timeout=60000)
        await page.wait_for_timeout(2000)

        # Get all "View more" links — group by candidate name, keep LAST per candidate
        html = await page.content()
        from bs4 import BeautifulSoup as _BS
        soup = _BS(html, "html.parser")

        # Each candidate card: <td><div class="details-name"><h4 class="bg-dark-blu">NAME</h4>...<a href="...">View more</a></td>
        # When no time filter: first occurrence wins (newest listed first)
        # When --before-time is set: collect ALL occurrences per name, process each individually
        seen = {}  # {candidate_name: href} — first occurrence wins (latest affidavit)
        seen_all = {}  # {candidate_name: [href1, href2, ...]} — all occurrences when time filter active
        all_links = soup.find_all("a", string=lambda x: x and "view more" in x.lower())
        logger.info(f"  {len(all_links)} total candidate entries found")

        for link in all_links:
            href = link.get("href", "")
            if not href:
                continue
            # Walk up to find the td/div containing this link, then find h4
            name = "Unknown"
            parent = link
            for _ in range(8):
                parent = parent.parent
                if not parent:
                    break
                h4 = parent.find("h4", class_="bg-dark-blu")
                if h4:
                    name = h4.get_text(strip=True)
                    break
            if name not in seen:  # first occurrence wins (newest is listed first)
                seen[name] = href
            # Also track all occurrences for time-filtered runs
            if name not in seen_all:
                seen_all[name] = []
            seen_all[name].append(href)

        # When time filter is active, use all occurrences (each will be checked against time filter)
        # BUT still deduplicate by name — keep only the LATEST entry within the time window per candidate
        if self.before_time:
            # Pass all hrefs but _process_candidate will handle time filtering
            # We still need name-based dedup AFTER time filtering
            # So pass all entries — the time filter in _process_candidate will skip out-of-window ones
            # and the dedup_key (candidateid) in tracker handles replacing old with new
            candidate_hrefs = [(name, href) for name, hrefs in seen_all.items() for href in hrefs]
            logger.info(f"  {len(seen_all)} unique candidates ({len(candidate_hrefs)} total entries, time filter active)")
        else:
            candidate_hrefs = list(seen.items())
            logger.info(f"  {len(candidate_hrefs)} unique candidates after dedup")

        new_total = 0
        processed_names_this_run = set()  # prevent downloading same name twice in one run
        for cand_name, href in candidate_hrefs:
            # Skip if we already successfully processed this candidate name in this run
            if cand_name in processed_names_this_run and cand_name != "Unknown":
                logger.info(f"  Skipping {cand_name} — already processed in this run")
                continue
            new_files = await self._process_candidate(context, href, out_dir, downloaded)
            # Mark as processed only if time filter passed (new_files >= 0)
            # If new_files == -1, time filter rejected — try next entry for same name
            if new_files >= 0:
                processed_names_this_run.add(cand_name)
            new_total += max(0, new_files)  # don't count -1
            if new_files > 0:
                save_tracker(tracker_file, downloaded)

        logger.info(f"  {const_name}: {new_total} new files downloaded")
        return new_total

    async def _process_candidate(self, context: BrowserContext, profile_url: str,
                                  out_dir: Path, downloaded: dict) -> int:
        """Open candidate profile, intercept download, save PDF."""
        new_files = 0
        download_info = {}  # captured from /increaseDownloadCount

        detail_page = await context.new_page()

        # Intercept the increaseDownloadCount call to get nomid+candidateid
        async def on_request(req):
            if "increaseDownloadCount" in req.url and req.post_data:
                params = {}
                for part in req.post_data.split("&"):
                    if "=" in part:
                        k, v = part.split("=", 1)
                        params[k] = v
                download_info.update(params)

        detail_page.on("request", on_request)

        try:
            await detail_page.goto(profile_url, wait_until="domcontentloaded", timeout=20000)
            await detail_page.wait_for_timeout(1000)

            # Get candidate name and upload timestamp from page
            name_el = await detail_page.query_selector("h3, h4, .candidate-name, [class*='name'], td:has-text('Name') + td")
            cand_name = (await name_el.inner_text()).strip() if name_el else "Candidate"

            # Parse "Affidavit Uploaded On" timestamp
            upload_time = ""
            try:
                from bs4 import BeautifulSoup
                html = await detail_page.content()
                soup = BeautifulSoup(html, "html.parser")
                for strong in soup.find_all("strong"):
                    if "Affidavit Uploaded On" in strong.get_text():
                        # Structure: <div class="items"><span><strong>label</strong></span> <span>VALUE</span></div>
                        items_div = strong.find_parent("div")
                        if items_div:
                            spans = items_div.find_all("span")
                            if len(spans) >= 2:
                                upload_time = spans[1].get_text(strip=True)
                        break
            except Exception:
                pass

            logger.info(f"  Processing: {cand_name} | Uploaded: {upload_time or 'unknown'}")

            # Filter by affidavit upload time if --before-time is set
            # upload_time format: "30th March, 2026 17:20:36"
            if self.before_time and upload_time:
                try:
                    from datetime import datetime as _dt
                    # Parse upload_time
                    for fmt in ["%dth %B, %Y %H:%M:%S", "%dst %B, %Y %H:%M:%S",
                                "%dnd %B, %Y %H:%M:%S", "%drd %B, %Y %H:%M:%S"]:
                        try:
                            # Remove ordinal suffix for parsing
                            import re as _re
                            clean = _re.sub(r'(\d+)(st|nd|rd|th)', r'\1', upload_time)
                            ut = _dt.strptime(clean, "%d %B, %Y %H:%M:%S")
                            break
                        except ValueError:
                            continue
                    else:
                        ut = None

                    if ut:
                        limit_h, limit_m = map(int, self.before_time.split(":"))
                        limit_time = ut.replace(hour=limit_h, minute=limit_m, second=0)
                        if ut > limit_time:
                            logger.info(f"  Skipping {cand_name} — uploaded at {upload_time}, after {self.before_time}")
                            return -1  # -1 = time filter rejected, try next entry for same name
                except Exception as e:
                    logger.warning(f"  Time filter parse error: {e}")

            # Filter by application uploaded date if specified
            if self.filter_date:
                app_uploaded = ""
                try:
                    for strong in soup.find_all("strong"):
                        if "Application Uploaded" in strong.get_text():
                            # Structure: <div class="col-sm-6"><label><p><strong>label</strong></p></label></div>
                            #            <div class="col-sm-6"><div><p>VALUE</p></div></div>
                            col_div = strong.find_parent("div", class_="col-sm-6")
                            if col_div:
                                nxt = col_div.find_next_sibling("div")
                                if nxt:
                                    app_uploaded = nxt.get_text(strip=True)
                            break
                except Exception:
                    pass

                from datetime import datetime as _dt
                try:
                    fd = _dt.strptime(self.filter_date, "%d-%m-%Y")
                    day = fd.day
                    suffix = "th" if 11 <= day <= 13 else {1:"st",2:"nd",3:"rd"}.get(day%10,"th")
                    date_match = f"{day}{suffix} {fd.strftime('%B')}, {fd.year}"
                    logger.info(f"  App uploaded: [{app_uploaded}] | Looking for: [{date_match}]")
                    if date_match not in app_uploaded:
                        logger.info(f"  Skipping {cand_name} — date mismatch")
                        return 0
                except Exception:
                    pass

            # Find the LATEST Download button — when multiple affidavits exist, pick highest ID
            # Structure: <button onclick="return increaseDownloadCount(5629);">Download</button>
            import re as _re2
            all_download_btns = await detail_page.query_selector_all(
                "a:has-text('Download'), button:has-text('Download')"
            )
            download_btn = None
            if len(all_download_btns) <= 1:
                download_btn = all_download_btns[0] if all_download_btns else None
            else:
                # Multiple affidavits — pick the one with highest increaseDownloadCount ID
                best_id = -1
                for btn in all_download_btns:
                    onclick = await btn.get_attribute("onclick") or ""
                    if not onclick:
                        # Check parent anchor
                        parent_a = await btn.evaluate_handle("el => el.closest('a') || el")
                        onclick = await parent_a.get_attribute("onclick") or ""
                    # Also check inner button
                    inner_btn = await btn.query_selector("button[onclick]")
                    if inner_btn:
                        onclick = await inner_btn.get_attribute("onclick") or onclick
                    m = _re2.search(r'increaseDownloadCount\((\d+)\)', onclick)
                    if m:
                        affid = int(m.group(1))
                        if affid > best_id:
                            best_id = affid
                            download_btn = btn
                    elif download_btn is None:
                        download_btn = btn
                if best_id > 0:
                    logger.info(f"  Multiple affidavits found — using latest (ID={best_id})")
            if not download_btn:
                logger.warning(f"  No download button for {cand_name}")
                return 0

            # Check dedup before downloading
            # We'll check after we get the nomid from the intercepted request
            try:
                async with detail_page.expect_download(timeout=15000) as dl_info:
                    await download_btn.click()
                dl = await dl_info.value

                # Dedup key = candidateid + upload_time
                # If same candidate uploads a new affidavit tomorrow, upload_time changes → new download
                nomid = download_info.get("nomid", "")
                candidateid = download_info.get("candidateid", "")
                dedup_key = candidateid if candidateid else dl.url

                # Build version string from upload time
                version_key = f"{dedup_key}|{upload_time}" if upload_time else dedup_key

                if version_key in downloaded:
                    logger.info(f"  Skipping {cand_name} — already have latest ({upload_time})")
                    await dl.cancel()
                    return 0

                # If same candidate exists but with older timestamp — delete old file
                if dedup_key in downloaded:
                    old_entry = downloaded[dedup_key]
                    old_filename = old_entry if isinstance(old_entry, str) else old_entry.get("filename", "")
                    old_path = out_dir / old_filename
                    if old_path.exists():
                        old_path.unlink()
                        logger.info(f"  Replacing old affidavit for {cand_name} (new upload: {upload_time})")

                # Use stable filename based on candidateid — not the ECI-generated timestamp name
                nomid = download_info.get("nomid", "")
                candidateid = download_info.get("candidateid", "")
                stable_name = f"Affidavit_{candidateid}.pdf" if candidateid else (dl.suggested_filename or f"Affidavit_{nomid}.pdf")
                save_path = out_dir / stable_name
                save_path.parent.mkdir(parents=True, exist_ok=True)
                await dl.save_as(save_path)

                size_kb = save_path.stat().st_size / 1024
                downloaded[version_key] = stable_name
                downloaded[dedup_key] = {"filename": stable_name, "upload_time": upload_time}
                new_files += 1
                logger.info(f"  Downloaded: {cand_name} → {stable_name} ({size_kb:.1f} KB)")

                # Upload to Control Plane if enabled
                if self._upload_client:
                    pdf_bytes = save_path.read_bytes()
                    await self._upload_client.upload_pdf(
                        constituency_name=out_dir.name,
                        filename=stable_name,
                        pdf_bytes=pdf_bytes,
                    )

            except Exception as e:
                logger.warning(f"  Download failed for {cand_name}: {e}")

        except Exception as e:
            logger.error(f"  Error on profile {profile_url}: {e}")
        finally:
            await detail_page.close()

        return new_files


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

async def main():
    parser = argparse.ArgumentParser(description="ECI Candidate Affidavit Scraper")
    parser.add_argument("--state", default="S22", help="State code (S22=Tamil Nadu)")
    parser.add_argument("--election", default="32-AC-GENERAL-3-60", help="Election type value")
    parser.add_argument("--constituency", default=None, nargs="+", help="Filter to one or more constituency names e.g. PERAMBUR ALANDUR")
    parser.add_argument("--upload", action="store_true", help="Upload PDFs to Control Plane API after download")
    parser.add_argument("--date", default=None, help="Only download candidates uploaded on this date e.g. 01-04-2026")
    parser.add_argument("--before-time", default=None, dest="before_time", help="Only download affidavits uploaded before this time e.g. 16:00")
    args = parser.parse_args()

    logger.info(f"ECI Affidavit Scraper | State: {args.state} | Election: {args.election}")
    if args.constituency:
        logger.info(f"Filtering to constituency: {args.constituency}")
    if args.date:
        logger.info(f"Filtering to uploaded date: {args.date}")

    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    scraper = AffidavitScraper(
        state_code=args.state,
        election_type=args.election,
        target_constituency=args.constituency,
        upload=args.upload,
        filter_date=args.date,
        before_time=args.before_time,
    )
    if args.constituency:
        logger.info(f"Filtering to constituencies: {args.constituency}")
    await scraper.run()
    logger.info("Done.")


if __name__ == "__main__":
    asyncio.run(main())